from odoo import models
import os
import sys
import logging

# Add the 'libs' directory to the system path dynamically
module_dir = os.path.dirname(os.path.abspath(__file__))
libs_dir = os.path.join(module_dir, 'libs')
#sys.path.append(libs_dir)
sys.path.append(r'C:\Program Files\Odoo 17.0e.20240415\server\addons\excel_filler\libs')

import openpyxl

class ExcelFiller(models.Model):
    _name = 'excel.filler'
    _description = 'Excel Filler'

    def fill_excel_file(self, file_path, data_array):
        """
        Fill the Excel file at the specified path with data from an array and return a status message.

        :param file_path: The path to the Excel file.
        :param data_array: A list of tuples where each tuple contains (sheet_name, row, col, value)
                           Example: [('Sheet1', 1, 1, 'Hello'), ('Sheet1', 2, 3, 123)]
        :return: A string message indicating the result of the operation.
        """
        if not os.path.isfile(file_path):
            return f"File not found: {file_path}"

        try:
            workbook = openpyxl.load_workbook(file_path)
        except Exception as e:
            return f"Failed to load the Excel file: {file_path}. Error: {e}"

        response_messages = []
        for entry in data_array:
            if len(entry) != 4:
                response_messages.append(f"Invalid data format: {entry}. Expected format: (sheet_name, row, col, value).")
                continue

            sheet_name, row, col, value = entry

            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                original_value = sheet.cell(row=row, column=col).value
                sheet.cell(row=row, column=col).value = value
                response_messages.append(
                    f"Updated cell ({row}, {col}) in sheet '{sheet_name}' from '{original_value}' to '{value}'."
                )
            else:
                response_messages.append(f"Sheet '{sheet_name}' does not exist in the workbook.")

        try:
            workbook.save(file_path)
            response_messages.append(f"Excel file saved successfully: {file_path}")
        except Exception as e:
            response_messages.append(f"Failed to save the Excel file: {file_path}. Error: {e}")

        return "\n".join(response_messages)
